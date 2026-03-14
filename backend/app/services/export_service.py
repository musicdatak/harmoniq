"""Export service — generate Excel and text from scheduled playlists."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate_excel(playlist_name: str, tracks: list[dict]) -> bytes:
    """Generate an .xlsx file from scheduled track data.

    tracks: list of dicts with keys: position, title, artist, key_camelot,
            key_musical, bpm, energy, transition_score, transition_label
    """
    wb = Workbook()
    ws = wb.active
    ws.title = playlist_name[:31]  # Excel sheet name max 31 chars

    headers = ["#", "Title", "Artist", "Key (Camelot)", "Key (Musical)", "BPM", "Energy", "Transition"]
    header_fill = PatternFill(start_color="00E5C7", end_color="00E5C7", fill_type="solid")
    header_font = Font(bold=True, color="0A0A0F")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, track in enumerate(tracks, 2):
        ws.cell(row=row_idx, column=1, value=track.get("position", row_idx - 1))
        ws.cell(row=row_idx, column=2, value=track.get("title", ""))
        ws.cell(row=row_idx, column=3, value=track.get("artist", ""))
        ws.cell(row=row_idx, column=4, value=track.get("key_camelot", ""))
        ws.cell(row=row_idx, column=5, value=track.get("key_musical", ""))
        bpm = track.get("bpm")
        ws.cell(row=row_idx, column=6, value=float(bpm) if bpm else None)
        energy = track.get("energy")
        ws.cell(row=row_idx, column=7, value=float(energy) if energy else None)

        label = track.get("transition_label", "")
        score = track.get("transition_score")
        transition = f"{label} ({float(score):.0f})" if label and score else ""
        ws.cell(row=row_idx, column=8, value=transition)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_text(playlist_name: str, tracks: list[dict]) -> str:
    """Generate numbered text list from scheduled track data."""
    lines = [f"# {playlist_name}", ""]
    for track in tracks:
        pos = track.get("position", "?")
        artist = track.get("artist", "")
        title = track.get("title", "")
        key = track.get("key_camelot", "")
        bpm = track.get("bpm", "")
        label = track.get("transition_label", "")

        line = f"{pos}. {artist} - {title}"
        meta_parts = []
        if key:
            meta_parts.append(key)
        if bpm:
            meta_parts.append(f"{float(bpm):.0f} BPM")
        if label:
            meta_parts.append(label)
        if meta_parts:
            line += f"  [{', '.join(meta_parts)}]"
        lines.append(line)

    return "\n".join(lines)
