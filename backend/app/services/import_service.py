"""Import service — parse Excel and text track lists."""

import io
import re

from openpyxl import load_workbook

# Fuzzy header matching
_TITLE_HEADERS = {"title", "song", "track", "titre", "chanson", "name", "song name", "track name", "song title"}
_ARTIST_HEADERS = {"artist", "performer", "artiste", "interprète", "interprete", "band", "artist name"}
_KEY_HEADERS = {"key", "tonalité", "tonalite", "musical key", "clé", "cle"}
_BPM_HEADERS = {"bpm", "tempo", "beats per minute"}
_ENERGY_HEADERS = {"energy", "énergie", "energie", "nrj"}

_TEXT_SEPARATORS = re.compile(r"\s+[-–—|]\s+|\t")


def _match_header(header: str, candidates: set[str]) -> bool:
    return header.strip().lower() in candidates


def _detect_columns(headers: list[str]) -> dict[str, int | None]:
    """Map logical columns to header indices via fuzzy matching."""
    mapping: dict[str, int | None] = {
        "title": None, "artist": None, "key": None, "bpm": None, "energy": None,
    }
    header_sets = {
        "title": _TITLE_HEADERS,
        "artist": _ARTIST_HEADERS,
        "key": _KEY_HEADERS,
        "bpm": _BPM_HEADERS,
        "energy": _ENERGY_HEADERS,
    }
    for idx, h in enumerate(headers):
        h_lower = h.strip().lower()
        for field, candidates in header_sets.items():
            if h_lower in candidates:
                mapping[field] = idx
                break
    return mapping


def parse_excel(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse an Excel/CSV file into a list of track dicts."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # First row = headers
    headers = [str(cell or "") for cell in rows[0]]
    col_map = _detect_columns(headers)

    if col_map["title"] is None or col_map["artist"] is None:
        # Fallback: assume col 0 = artist, col 1 = title
        if len(headers) >= 2:
            col_map["artist"] = 0
            col_map["title"] = 1

    if col_map["title"] is None or col_map["artist"] is None:
        return []

    tracks = []
    for row in rows[1:]:
        cells = list(row)
        title_val = cells[col_map["title"]] if col_map["title"] < len(cells) else None
        artist_val = cells[col_map["artist"]] if col_map["artist"] < len(cells) else None

        if not title_val or not artist_val:
            continue

        track: dict = {
            "title": str(title_val).strip(),
            "artist": str(artist_val).strip(),
        }

        if col_map["key"] is not None and col_map["key"] < len(cells) and cells[col_map["key"]]:
            track["key"] = str(cells[col_map["key"]]).strip()
        if col_map["bpm"] is not None and col_map["bpm"] < len(cells) and cells[col_map["bpm"]]:
            try:
                track["bpm"] = float(cells[col_map["bpm"]])
            except (ValueError, TypeError):
                pass
        if col_map["energy"] is not None and col_map["energy"] < len(cells) and cells[col_map["energy"]]:
            try:
                track["energy"] = float(cells[col_map["energy"]])
            except (ValueError, TypeError):
                pass

        tracks.append(track)

    return tracks


def parse_text(text: str) -> list[dict]:
    """Parse text lines into track dicts. Splits on ' - ', ' – ', ' — ', ' | ', or tab."""
    tracks = []
    for line in text.strip().splitlines():
        line = line.strip()
        # Strip leading numbering like "1. " or "1) "
        line = re.sub(r"^\d+[\.\)]\s*", "", line)
        if not line:
            continue

        parts = _TEXT_SEPARATORS.split(line, maxsplit=1)
        if len(parts) == 2:
            artist = parts[0].strip()
            title = parts[1].strip()
            if artist and title:
                tracks.append({"artist": artist, "title": title})

    return tracks
